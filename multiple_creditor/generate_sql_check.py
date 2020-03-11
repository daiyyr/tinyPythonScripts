from openpyxl import load_workbook
import os

#Step 1: Execute open excel workbook 
ACTION1_File = load_workbook(r"C:\Users\Teemo\Downloads\duplicate_creditors_-_blank_bank_acc_and_gst_-_same_name.xlsx", )
ACTION1_File_Sheet = ACTION1_File.get_sheet_names()[0]  
ACTION1_File_Sheet_Name = ACTION1_File.get_sheet_by_name(ACTION1_File_Sheet)

directory = os.path.dirname(os.path.realpath(__file__))
sql_file = r"select_multiple_cinvoice_16042019.sql"
skip_file = r"skip_finalized_id_list.txt"
with open(skip_file, 'w+') as fp:
        fp.write("")
with open(sql_file, 'w+') as fp:
        fp.write("""
drop TEMPORARY TABLE IF EXISTS duplicate_cinvoice_temp;
CREATE TEMPORARY TABLE IF NOT EXISTS duplicate_cinvoice_temp AS 
(SELECT ''as group_id, '' as cinvoice_id, cinvoice_num, '' as creditor_master_name, '' as bodycorp_code FROM cinvoices WHERE 1=2);""")

line_count = 0
finalized_count = 0

to_be_deleted_creditor_id_12032019 = ""
finalized_creditor_id_12032019 = ""
all_finalized_code = ""

for line in ACTION1_File_Sheet_Name.iter_rows():
        line_count += 1
        if line_count < 3:      #change this number every time when excel updated
                continue
        cell_count = 0
        for c in line:
                cell_count += 1
                if cell_count == 1:
                        if c.value != None:     #find new group, so finish last group, save result in file
                                if to_be_deleted_creditor_id_12032019 != "" and finalized_creditor_id_12032019 != "":
                                        with open(os.path.join(directory, sql_file), 'a+') as fp:
                                                fp.write(
                                                        """
drop TEMPORARY TABLE IF EXISTS to_be_deleted_creditor_id_12032019;
CREATE TEMPORARY TABLE IF NOT EXISTS to_be_deleted_creditor_id_12032019 AS (SELECT creditor_master_id FROM creditor_master WHERE 
"""+to_be_deleted_creditor_id_12032019+"""

);
drop TEMPORARY TABLE IF EXISTS finalized_creditor_id_12032019;
CREATE TEMPORARY TABLE IF NOT EXISTS finalized_creditor_id_12032019 AS (SELECT creditor_master_id FROM creditor_master WHERE creditor_master_id = 
'"""+finalized_creditor_id_12032019+"""'

);
insert into duplicate_cinvoice_temp (
    select group_id,cinvoice_id,cinvoice_num,creditor_master_name,bodycorp_code from (
        select """+str(finalized_count)+""" as group_id, 
        group_concat(cinvoice_id SEPARATOR ', ') as cinvoice_id,
        cinvoice_num,
        group_concat(creditor_master_name separator ', ') as creditor_master_name,
        group_concat(bodycorp_code separator ', ') as bodycorp_code,
        count(*) as cc
        from cinvoices
        join bodycorps on bodycorp_id = cinvoice_bodycorp_id
        join creditor_master on creditor_master_id = cinvoice_creditor_id
        where cinvoice_creditor_id = (SELECT creditor_master_id FROM finalized_creditor_id_12032019) or cinvoice_creditor_id in (SELECT creditor_master_id FROM to_be_deleted_creditor_id_12032019 )
        group by cinvoice_num
    ) zzz where cc > 1
);
"""
                                                )
                                
                                to_be_deleted_creditor_id_12032019 = ""
                                finalized_creditor_id_12032019 = ""

                if cell_count == 2:
                        if c.value != None:
                                if c.fill.fgColor.index == 9: # 9 -> green; '00000000' -> white; 'FF7030A0' -> purple
                                        if finalized_creditor_id_12032019 == "":
                                                finalized_creditor_id_12032019 = str(c.value)
                                                all_finalized_code += "'" + str(c.value) + "'," 
                                        #else:
                                                #raise ValueError('find more than one finalized creditor under one label: {}'.format(str(c.value)))
                                                #if there are more than one finalized creditor under one label, leave it
                                                #Do not merge it. Merge others to the first finalized one

                                        finalized_count += 1
                                else:
                                        if to_be_deleted_creditor_id_12032019 == "":
                                                to_be_deleted_creditor_id_12032019 += " creditor_master_id = '"+str(c.value)+"' "
                                        else:
                                                to_be_deleted_creditor_id_12032019 += " or creditor_master_id = '"+str(c.value)+"' "
                                break

with open(os.path.join(directory, sql_file), 'a+') as fp:
    fp.write("""
select * from duplicate_cinvoice_temp;
""")

print("{} finalized creditors were found. Execute sql to find duplicate cinvoices".format(finalized_count))
print(all_finalized_code)