from openpyxl import load_workbook
import os

#Step 1: Execute open excel workbook 
ACTION1_File = load_workbook(r"C:\Users\Teemo\Downloads\duplicate_creditors_-_blank_bank_acc_and_gst_-_same_name.xlsx", )
ACTION1_File_Sheet = ACTION1_File.get_sheet_names()[0]  
ACTION1_File_Sheet_Name = ACTION1_File.get_sheet_by_name(ACTION1_File_Sheet)

directory = os.path.dirname(os.path.realpath(__file__))
sql_file = r"delete_multiple_creditor_17042019.sql"
with open(sql_file, 'w+') as fp:
        fp.write("""
        -- !!!!!
        -- !!!!!
        -- change 26041 to a temperary testing creditor master id before execute
        -- !!!!!
        -- !!!!!
        """)

line_count = 0
finalized_count = 0

to_be_deleted_creditor_id_12032019 = ""
finalized_creditor_id_12032019 = ""
all_finalized_code = ""
update_code = ""
current_creditor_name = ""

for line in ACTION1_File_Sheet_Name.iter_rows():
        creditor_id_of_line = ""
        line_count += 1
        if line_count < 2:      #change this number every time when excel updated
                continue
        creditor_name = line[3].value
        if current_creditor_name=="" or (creditor_name is not None and current_creditor_name is not None and current_creditor_name.upper().replace(" ","")==creditor_name.upper().replace(" ","") ):   #first line or still in old group
            current_creditor_name = creditor_name
        else:   #find new group, so finish last group, save result in file
            current_creditor_name = creditor_name
            if to_be_deleted_creditor_id_12032019 != "" and finalized_creditor_id_12032019 != "":
                with open(os.path.join(directory, sql_file), 'a+') as fp:
                        fp.write(
                                                        """
drop TEMPORARY TABLE IF EXISTS to_be_deleted_creditor_id_12032019;
CREATE TEMPORARY TABLE IF NOT EXISTS to_be_deleted_creditor_id_12032019 AS (SELECT creditor_master_id FROM creditor_master WHERE 
"""+to_be_deleted_creditor_id_12032019+"""

);
drop TEMPORARY TABLE IF EXISTS finalized_creditor_id_12032019;
CREATE TEMPORARY TABLE IF NOT EXISTS finalized_creditor_id_12032019 AS (SELECT creditor_master_id FROM creditor_master WHERE creditor_master_id = 
'"""+finalized_creditor_id_12032019+"""'

);
UPDATE cinvoices SET cinvoice_creditor_id = (SELECT creditor_master_id FROM finalized_creditor_id_12032019) WHERE cinvoice_creditor_id IN (SELECT creditor_master_id FROM to_be_deleted_creditor_id_12032019 );
UPDATE cpayments SET cpayment_creditor_id = (SELECT creditor_master_id FROM finalized_creditor_id_12032019) WHERE cpayment_creditor_id IN (SELECT creditor_master_id FROM to_be_deleted_creditor_id_12032019 );
insert into creditor_history_comms (creditor_history_comm_creditor_id,creditor_history_comm_comm_id) 
    select finalized_creditor_id_12032019.creditor_master_id,creditor_comm_comm_id 
    FROM finalized_creditor_id_12032019 
    join to_be_deleted_creditor_id_12032019 on 1=1 
    join creditor_comms on creditor_comm_creditor_id = to_be_deleted_creditor_id_12032019.creditor_master_id;
delete from creditor_comms where creditor_comm_creditor_id in (SELECT creditor_master_id FROM to_be_deleted_creditor_id_12032019);
update gl_transactions set gl_transaction_creditor_id = (SELECT creditor_master_id FROM finalized_creditor_id_12032019) where gl_transaction_creditor_id IN (SELECT creditor_master_id FROM to_be_deleted_creditor_id_12032019 );
delete from creditor_master where creditor_master_id in (SELECT creditor_master_id FROM to_be_deleted_creditor_id_12032019 );
                                                        """ + update_code
                                                )
                                
            to_be_deleted_creditor_id_12032019 = ""
            finalized_creditor_id_12032019 = ""
            update_code = ""
        cell_count = 0
        for c in line:
                cell_count += 1
                if cell_count == 1:
                        if c.value != None:
                                creditor_id_of_line = c.value
                                if c.fill.fgColor.index == 9: # 9 -> green; '00000000' -> white; 'FF7030A0' -> purple
                                        if finalized_creditor_id_12032019 == "":
                                                finalized_creditor_id_12032019 = str(c.value)
                                                all_finalized_code += "'" + str(c.value) + "'," 

                                        # else:
                                        #         raise ValueError('find more than one finalized creditor under one label: {}'.format(str(c.value)))
                                        #if there are more than one finalized creditor under one label, leave it
                                        #Do not merge it. Merge others to the first finalized one

                                        finalized_count += 1
                                else:
                                        if to_be_deleted_creditor_id_12032019 == "":
                                                to_be_deleted_creditor_id_12032019 += " creditor_master_id = '"+str(c.value)+"' "
                                        else:
                                                to_be_deleted_creditor_id_12032019 += " or creditor_master_id = '"+str(c.value)+"' "
                if cell_count == 7:
                        if c.value != None:
                                update_code += """
update creditor_master set creditor_master_code = '"""+str(c.value)+"""' where creditor_master_id = """ + str(creditor_id_of_line) + """;
                                                """
                        break

print("{} finalized creditors were taken care of.".format(finalized_count))
print(all_finalized_code)